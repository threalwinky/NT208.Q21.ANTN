from __future__ import annotations

import argparse
import asyncio
import csv
import json
import re
import shutil
import subprocess
from dataclasses import dataclass
from datetime import datetime, timezone
from io import StringIO
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urldefrag, urljoin, urlparse, urlunparse

import httpx
from bs4 import BeautifulSoup

from app.services.data_paths import resolve_data_dir
from app.services.text_utils import clean_text, content_hash

BLOCKED_EXTENSIONS = {
    ".gif",
    ".webp",
    ".svg",
    ".zip",
    ".rar",
    ".mp4",
    ".mp3",
    ".doc",
    ".docx",
    ".ppt",
    ".pptx",
}

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
SPREADSHEET_EXTENSIONS = {".csv", ".xls", ".xlsx"}

BLOCKED_KEYWORDS = [
    "tuyen-sinh",
    "admission",
    "admissions",
    "mailto:",
    "tel:",
    "facebook.com",
    "youtube.com",
    "/en/",
    "/en_us/",
    "lang=en",
    "module-variations",
    "/icons",
    "/logo-",
]

GLOBAL_PRIORITY_KEYWORDS = [
    "sinh-vien",
    "student",
    "hoc-vu",
    "dao-tao",
    "thong-bao",
    "huong-dan",
    "quy-che",
    "quy-dinh",
    "tot-nghiep",
    "hoc-phi",
    "hoc-bong",
    "giay",
    "lich-thi",
    "lich-hoc",
    "chuong-trinh",
    "chuan-dau-ra",
    "ngoai-ngu",
    "tieng-anh",
    "chung-chi",
    "vnu-ept",
    "toeic",
    "ielts",
    "vstep",
    "nganh",
    "hoc-phan",
    "mon-hoc",
    "hoc-ky",
    "khoa-luan",
    "de-cuong",
    "thuc-tap",
    "huong-nghiep",
    "giai-thuong",
    "doan-hoi",
    "cau-lac-bo",
    "clb",
]

SOURCE_SEED_PATHS: dict[str, list[str]] = {
    "uit-main": [
        "/",
        "/dao-tao",
        "/sinh-vien",
        "/thong-bao",
        "/tin-tuc-su-kien",
    ],
    "uit-daa": [
        "/sinhvien",
        "/thongbao",
        "/huong-dan",
        "/bieu-mau",
        "/quy-trinh",
        "/kehoachnam",
    ],
    "uit-student": [
        "/",
        "/thongbao",
        "/quydinhdaotao",
        "/quydinh",
        "/quy-che-quy-dinh-dao-tao-dai-hoc-cua-dhqg-hcm",
        "/moc-thoi-gian",
        "/hoc-bong",
    ],
    "uit-ctsv": [
        "/",
        "/quy-che-quy-dinh",
        "/hoc-bong",
        "/tin-tuc",
        "/ho-tro-sinh-vien",
    ],
    "uit-khtc": [
        "/",
        "/thong-bao",
        "/sinh-vien",
        "/thanh-toan-tai-chinh",
    ],
    "uit-courses": [
        "/",
        "/login/index.php",
    ],
    "uit-oep": [
        "/vi",
        "/vi/tong-quan",
        "/vi/gallery-category/hoc-tap",
        "/vi/chuong-trinh-tien-tien",
        "/vi/tong-quan-ve-chuong-trinh-chat-luong-cao",
        "/vi/tong-quan-ve-chuong-trinh-lien-ket-uon",
        "/vi/tong-quan-ve-chuong-trinh-lien-ket-bcu",
        "/vi/cu-nhan-tai-nang-thiet-ke-vi-mach",
    ],
    "uit-se": [
        "/",
        "/vi/tin-tuc/thong-bao-hoc-vu.html",
    ],
    "uit-fit": [
        "/",
        "/index.php/dao-tao/bac-dai-hoc/thong-bao-he-dai-hoc",
        "/index.php/dao-tao/bac-dai-hoc/chuong-trinh-dao-dai-hoc",
        "/index.php/dao-tao/bac-dai-hoc/cac-quy-trinh-bieu-mau-sinh-vien",
        "/index.php/tin-tuc/tin-h-c-v/thong-bao-cac-h-dao-t-o/thong-bao-he-dai-hoc",
    ],
    "uit-cs": [
        "/",
        "/category/tin-tuc/thong-bao/",
        "/category/tin-tuc/sinh-vien/",
        "/category/tin-tuc/hoat-dong-doan-hoi-clb/",
        "/category/luu-tru/quy-dinh/",
        "/category/luu-tru/bieu-mau/",
        "/category/tin-tuc/huong-nghiep/",
        "/he-dai-tra/",
        "/he-tai-nang/",
        "/bcu/",
    ],
    "uit-httt": [
        "/",
        "/category/thong-bao-hoc-vu/",
        "/category/thong-tin-doan-hoi/",
        "/category/research/",
        "/category/viec-lam-thuc-tap/",
        "/category/seminar-cuoc-thi/",
        "/category/li%cc%a3ch-tru%cc%a3c-cvht/",
    ],
    "uit-nc": [
        "/",
        "/giao-vu/thong-bao-v-v-dang-ky-thuc-tap-doanh-nghiep-hk-2-nh-2025-2026.html",
        "/giao-vu/ke-hoach-bao-cao-thuc-tap-doanh-nghiep-hk-1-nam-hoc-2025-2026.html",
        "/tin-tuc/thong-bao-hoc-vu/quy-trinh-thuc-tap-2019.html",
        "/tin-tuc/doan-hoi",
        "/doan-khoa",
        "/lien-chi-hoi",
        "/cong-doan",
    ],
    "uit-fce": [
        "/",
        "/category/hoat-dong-doan-hoi/",
        "/category/cau-lac-bo/clb-ceec/",
        "/category/cau-lac-bo/clb-icc/",
        "/thong-bao-to-chuc-hoi-nghi-sinh-vien-khoa-ky-thuat-may-tinh-nam-hoc-2025-2026/",
    ],
}

SOURCE_DIRECT_URLS: dict[str, list[str]] = {
    "uit-student": [
        "https://student.uit.edu.vn/sites/default/files/files/QD%20ban%20hanh%20quy%20che%20dao%20tao%282%29.pdf",
        "https://student.uit.edu.vn/sites/default/files/files/Quy%20che%20Dao%20tao%20theo%20tin%20chi_UIT.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/1195-qd-dhqg_27-9-2019_quy_dinh_dao_tao_song_nganh_dhqg.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/1342-qd-dhqg_30-9-2022_quy_che_dao_tao.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/540_qd_dhqg_9-5-2023_quy_dinh_mo_nganh.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/170_qd_dhqg_27_02_2018_cong_nhan_chung_chi_ngoai_ngu.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/707_qd_dhqg23_6_2022_quy_che_tuyen_sinh_trinh_do_dh.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/836_qd_dhqg_19_7_2021_quy_che_tuyen_sinh_trinh_do_dai_hoc_dhqghcm.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/1658-qd-dhqg_24-12-2020_phe_duyet_thi_diem_bo_pcnl_svtn_dhqghcm.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/bb1528_dhqg_2-8-2022_hop_ve_cong_tac_dao_tao.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202310/cv671-dhqg_18-4-2022_su_dung_chung_chi_vnu-ept.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202412/507-qd-dhcntt-27-5-2024_quy_che_dao_tao_cho_sinh_vien.pdf",
        "https://student.uit.edu.vn/sites/daa/files/202505/1314-qd-bgddt_13-05-2025_ban_hanh_chuan_ctdt_ve_vi_mach_ban_dan_trinh_do_dh_ths.pdf",
        "https://student.uit.edu.vn/cqui/ctdt-khoa-2025",
        "https://student.uit.edu.vn/mot-so-quy-trinh-danh-cho-sinh-vien",
        "https://student.uit.edu.vn/thong-bao-quy-trinh-dang-ky-giay-gioi-thieu",
        "https://student.uit.edu.vn/thong-bao-lich-dkhp-va-tkb-du-kien-hk2-nam-hoc-2025-2026",
        "https://student.uit.edu.vn/thong-bao-thu-hoc-phi-hoc-ky-2-nam-hoc-2025-2026-trinh-do-dai-hoc",
        "https://student.uit.edu.vn/thong-bao-ve-viec-thu-hoc-phi-hoc-ky-1-nh-2025-2026-trinh-do-dai-hoc",
        "https://student.uit.edu.vn/ke-hoach-xet-tot-nghiep-dot-01-nam-2026",
        "https://student.uit.edu.vn/ke-hoach-xet-tot-nghiep-dot-04-nam-2025",
        "https://student.uit.edu.vn/thong-bao-nhan-don-nhap-hoc-lai-bao-luu-chuyen-nganh-song-nganh-thoi-hoc-hk-1-2025-2026",
        "https://student.uit.edu.vn/38-quy-trinh-xu-ly-hoc-vu-sinh-vien-dai-hoc-chinh-quy",
        "https://student.uit.edu.vn/32-quy-trinh-canh-bao-sinh-vien-he-dai-hoc-chinh-quy-ve-ket-qua-dang-ky-hoc-phan-va-ket-qua-hoc-tap",
    ],
    "uit-ctsv": [
        "https://ctsv.uit.edu.vn/sites/default/files/201607/33_qd_dhcntt_ctsv5_7_2016_scan.pdf",
        "https://ctsv.uit.edu.vn/sites/default/files/202204/47_tb_dhcntt19_4_2022_scan.pdf",
    ],
    "uit-khtc": [
        "https://khtc.uit.edu.vn/node/455",
        "https://khtc.uit.edu.vn/node/463",
        "https://khtc.uit.edu.vn/node/468",
        "https://khtc.uit.edu.vn/content/2025-2026-thong-bao-thu-hoc-phi-hk2-nh-2025-2026-trinh-do-dtdh",
        "https://khtc.uit.edu.vn/content/2025-2026-thong-bao-thu-hoc-phi-hk2-nh-2025-2026-trinh-do-dtdh-chuong-trinh-bcu",
        "https://khtc.uit.edu.vn/content/2025-2026-thong-bao-ve-viec-thu-hoc-phi-hk1-2025-2026ct-bcu-vb2cq-ltdh-song-nganh",
        "https://khtc.uit.edu.vn/sites/default/files/%5Bcurrent-date%3Acustom%3AYm%5D/quyet_dinh_muc_thu_hp_2025-2026_cqdtcntnkstn.pdf",
        "https://khtc.uit.edu.vn/sites/default/files/%5Bcurrent-date%3Acustom%3AYm%5D/quyet_dinh_muc_thu_hp_2025-2026_cttt.pdf",
        "https://khtc.uit.edu.vn/sites/default/files/%5Bcurrent-date%3Acustom%3AYm%5D/quyet_dinh_muc_thu_hp_2025-2026_clc.pdf",
        "https://khtc.uit.edu.vn/sites/default/files/202601/thong_bao_thu_hp_hk2_2025-2026.pdf",
        "https://khtc.uit.edu.vn/sites/default/files/%5Bcurrent-date%3Acustom%3AYm%5D/thong_bao_thu_hoc_phi_hk1_nam_hoc_2025-2026.pdf",
        "https://khtc.uit.edu.vn/sites/default/files/%5Bcurrent-date%3Acustom%3AYm%5D/thong_bao_thu_hoc_phi_hk1_2025-2026_k2025.pdf",
    ],
    "uit-oep": [
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202007/thong_bao_dkhp-hk1_2020-2021.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202112/thong_bao_dkhp-hk2_2021-2022.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202107/434-qd-dhcntt07-07-2021.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202512/1451-qd-dhcntt_26-11-2025.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202509/1008-qd-dhcntt_26-08-2025.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/201705/163-qd-dhcntt4-4-17_scan.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202107/435-qd-dhcntt07-07-2021.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/201904/211-qd-dhcntt17-4-2019_-_scan.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202206/131-qd-dhcntt_ngay_08-3-2022.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202512/1032-qd-dhcntt_03-9-2025.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202209/thong_bao_su_dung_chung_nhan_vnu-ept.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/201911/547_qd-dhcntt_30-8-2019.scan_.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202107/141-qd-dhcntt16-3-2018scan.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202210/828_qd-dhcntt_04-10-2022_quy_dinh_dao_tao_ngoai_ngu_doi_voi_he_chinh_qui_khoa_2022.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202404/364-qd-dhcntt_22-04-2024_sua_doi_quy_dinh_dao_tao_ngoai_ngu.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202406/560-qd-dhcntt_5-6-2024_sua_doi_quy_dinh_dao_tao_ngoai_ngu_doi_voi_he_dai_hoc_chinh_quy.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202107/546-qd-dhcntt_30-8-2019_quy_che_dao_tao_theo_hoc_che_tin_cho_he_dai_hoc_chinh_quy.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202107/148-qd-dhcntt-12-03-2021_chinh_sua_quy_che_dao_tao.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202210/790_qd-dhcntt_28-9-2022_.pdf",
        "https://oep.uit.edu.vn/sites/oep/files/uploads/files/202409/1393-qd-dhcntt_29-12-2023_cap_nhat_quy_che_dao_tao_theo_hoc_che_tin_chi_cho_he_dai_hoc_chinh_quy.pdf",
    ],
    "uit-daa": [
        "https://daa.uit.edu.vn/content/huong-dan-sinh-vien-dai-hoc-he-chinh-quy-thuc-hien-cac-quy-dinh-ve-chuan-qua-trinh-va-chuan",
        "https://daa.uit.edu.vn/content/chuong-trinh-dao-tao-tu-khoa-7-tro-di",
        "https://daa.uit.edu.vn/content/cu-nhan-nganh-ky-thuat-phan-mem-ap-dung-tu-khoa-19-2024",
        "https://daa.uit.edu.vn/content/cu-nhan-nganh-truyen-thong-da-phuong-tien-ap-dung-tu-khoa-20-2025",
        "https://daa.uit.edu.vn/mot-so-quy-trinh-danh-cho-sinh-vien",
        "https://daa.uit.edu.vn/kehoachnam",
        "https://daa.uit.edu.vn/sites/daa/files/uploads/pdtdh_bieu_do_ke_hoach_dao_tao_nam_hoc_2026-2027-ver-5.jpg",
        "https://daa.uit.edu.vn/sites/daa/files/uploads/pdtdh_bieu-do-ke-hoach-dao-tao-nam-hoc-2025-2026-ver-4-2.jpg",
    ],
    "uit-se": [
        "https://se.uit.edu.vn/vi/tin-tuc/thong-bao-hoc-vu.html",
        "https://se.uit.edu.vn/vi/tin-t%E1%BB%A9c/10-th%C3%B4ng-b%C3%A1o-h%E1%BB%8Dc-v%E1%BB%A5/2244-th%C3%B4ng-b%C3%A1o-c%E1%BA%ADp-nh%E1%BA%ADt-th%C3%B4ng-tin-th%E1%BB%B1c-t%E1%BA%ADp-doanh-nghi%E1%BB%87p-h%E1%BB%8Dc-k%E1%BB%B3-2-n%C4%83m-h%E1%BB%8Dc-2025-2026.html",
    ],
    "uit-fit": [
        "https://fit.uit.edu.vn/index.php/dao-tao/bac-dai-hoc/chuong-trinh-dao-dai-hoc",
        "https://fit.uit.edu.vn/index.php/dao-tao/bac-dai-hoc/cac-quy-trinh-bieu-mau-sinh-vien",
        "https://fit.uit.edu.vn/index.php/dao-tao/bac-dai-hoc/menu-cntt/chuan-dau-ra-cntt",
        "https://fit.uit.edu.vn/index.php/dao-tao/bac-dai-hoc/nganh-khoa-h-c-d-li-u/chuan-dau-ra-ds",
    ],
    "uit-cs": [
        "https://cs.uit.edu.vn/category/tin-tuc/thong-bao/",
        "https://cs.uit.edu.vn/category/tin-tuc/sinh-vien/",
        "https://cs.uit.edu.vn/category/tin-tuc/hoat-dong-doan-hoi-clb/",
        "https://cs.uit.edu.vn/category/luu-tru/quy-dinh/",
        "https://cs.uit.edu.vn/he-dai-tra/",
        "https://cs.uit.edu.vn/he-tai-nang/",
        "https://cs.uit.edu.vn/thong-bao-nop-de-cuong-khoa-luan-tot-nghiep-hk1-2025-2026/",
        "https://cs.uit.edu.vn/ke-hoach-xet-tot-nghiep-dot-03-nam-2025/",
    ],
    "uit-httt": [
        "https://httt.uit.edu.vn/category/thong-bao-hoc-vu/",
        "https://httt.uit.edu.vn/category/thong-tin-doan-hoi/",
        "https://httt.uit.edu.vn/category/research/",
        "https://httt.uit.edu.vn/category/viec-lam-thuc-tap/",
        "https://httt.uit.edu.vn/tong-ket-hoi-nghi-sinh-vien-nam-hoc-2025-2026-khoa-he-thong-thong-tin/",
        "https://httt.uit.edu.vn/hoi-nghi-sinh-vien-khoa-he-thong-thong-tin-nam-hoc-2025-2026/",
        "https://httt.uit.edu.vn/workshop-erp-va-chuyen-doi-so-trong-doanh-nghiep-co-hoi-cho-sinh-vien/",
    ],
    "uit-nc": [
        "https://nc.uit.edu.vn/giao-vu/thong-bao-v-v-dang-ky-thuc-tap-doanh-nghiep-hk-2-nh-2025-2026.html",
        "https://nc.uit.edu.vn/giao-vu/ke-hoach-bao-cao-thuc-tap-doanh-nghiep-hk-1-nam-hoc-2025-2026.html",
        "https://nc.uit.edu.vn/tin-tuc/thong-bao-hoc-vu/quy-trinh-thuc-tap-2019.html",
        "https://nc.uit.edu.vn/activities/thong-bao-khoi-dong-xet-chon-danh-hieu-sinh-vien-5-tot-cap-khoa-nam-hoc-2024-2025.html",
        "https://nc.uit.edu.vn/hoat-dong/hoi-nghi-sinh-vien-khoa-mang-may-tinh-va-truyen-thong-nam-2026.html",
        "https://nc.uit.edu.vn/doan-khoa",
        "https://nc.uit.edu.vn/lien-chi-hoi",
    ],
    "uit-fce": [
        "https://fce.uit.edu.vn/category/hoat-dong-doan-hoi/",
        "https://fce.uit.edu.vn/category/cau-lac-bo/clb-ceec/",
        "https://fce.uit.edu.vn/category/cau-lac-bo/clb-icc/",
        "https://fce.uit.edu.vn/thong-bao-to-chuc-hoi-nghi-sinh-vien-khoa-ky-thuat-may-tinh-nam-hoc-2025-2026/",
        "https://daa.uit.edu.vn/content/cu-nhan-nganh-ky-thuat-may-tinh-ap-dung-tu-khoa-18-2023",
        "https://daa.uit.edu.vn/content/cu-nhan-nganh-thiet-ke-vi-mach-ap-dung-tu-khoa-19-2024",
    ],
}

SKIP_STORE_PATHS = {
    "",
    "/",
    "/thong-bao",
    "/thongbao",
    "/tin-tuc-su-kien",
    "/sinhvien",
    "/ho-tro-sinh-vien",
    "/quy-che-quy-dinh",
}

SKIP_STORE_TITLES = {
    "module variations",
    "icons",
    "thông báo hệ đại học",
    "giới thiệu chung",
}

LOW_VALUE_TITLE_MARKERS = {
    "logo ",
    "logo các",
    "khoa khoa học và kỹ thuật thông tin",
}

UNCATEGORIZED_KEEP_MARKERS = {
    "chuan-dau-ra",
    "chuẩn đầu ra",
    "mon-hoc",
    "môn học",
    "hoc-phan",
    "học phần",
    "de-cuong",
    "đề cương",
    "khoa-luan",
    "khóa luận",
    "kltn",
    "tot-nghiep",
    "tốt nghiệp",
    "thuc-tap",
    "thực tập",
    "internship",
    "hoc-vu",
    "học vụ",
    "hoc-bong",
    "học bổng",
    "doan-hoi",
    "đoàn hội",
    "cau-lac-bo",
    "câu lạc bộ",
    "clb",
    "nghien-cuu",
    "nghiên cứu",
}

UNCATEGORIZED_BLOCK_MARKERS = {
    "module-variations",
    "icons",
    "logo",
    "gioi-thieu",
    "giới thiệu",
    "tuyen-sinh",
    "tuyển sinh",
    "lienthong",
    "liên thông",
    "cao-hoc",
    "cao học",
    "tuyen-dung",
    "tuyển dụng",
}

TRAINING_DISABLED_SOURCE_IDS = {"uit-courses"}

STUDENT_CRITICAL_SOURCE_IDS = {
    "uit-main",
    "uit-daa",
    "uit-student",
    "uit-ctsv",
    "uit-khtc",
    "uit-oep",
}

FACULTY_SOURCE_IDS = {
    "uit-se",
    "uit-fit",
    "uit-cs",
    "uit-httt",
    "uit-nc",
    "uit-fce",
}

HTML_TAIL_MARKERS = (
    "Bài viết liên quan",
    "PHÒNG ĐÀO TẠO ĐẠI HỌC",
    "@2012 Phòng Kế hoạch - Tài chính",
    "Trang 1 2 3 4 5 6 7 8 9",
)


@dataclass(slots=True)
class UrlCandidate:
    url: str
    score: int
    kind: str


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", value).strip("-").lower()
    return slug[:120] or "document"


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def normalize_url(url: str) -> str:
    clean_url = urldefrag(url.strip())[0]
    parsed = urlparse(clean_url)
    filtered_query = [
        (key, value)
        for key, value in parse_qsl(parsed.query, keep_blank_values=True)
        if key.lower()
        not in {
            "page",
            "start",
            "limitstart",
            "amp",
            "replytocom",
            "fbclid",
            "gclid",
            "utm_source",
            "utm_medium",
            "utm_campaign",
            "utm_term",
            "utm_content",
        }
    ]
    normalized = urlunparse(parsed._replace(query=urlencode(filtered_query, doseq=True)))
    return normalized.rstrip("/") or normalized


def is_pdf_url(url: str) -> bool:
    parsed = urlparse(url)
    return parsed.path.lower().endswith(".pdf") or ".pdf?" in url.lower()


def is_image_url(url: str) -> bool:
    return Path(urlparse(url).path.lower()).suffix in IMAGE_EXTENSIONS


def is_spreadsheet_url(url: str) -> bool:
    return Path(urlparse(url).path.lower()).suffix in SPREADSHEET_EXTENSIONS


def host_matches_domain(netloc: str, domain: str) -> bool:
    normalized_host = netloc.lower()
    normalized_domain = domain.lower()
    return normalized_host in {normalized_domain, f"www.{normalized_domain}"} or (
        normalized_domain.startswith("www.") and normalized_host == normalized_domain.removeprefix("www.")
    )


def is_allowed_url(url: str, domain: str) -> bool:
    parsed = urlparse(url)
    if not parsed.scheme.startswith("http"):
        return False
    if not host_matches_domain(parsed.netloc, domain):
        return False
    normalized = url.lower()
    if any(keyword in normalized for keyword in BLOCKED_KEYWORDS):
        return False
    if any(parsed.path.lower().endswith(extension) for extension in BLOCKED_EXTENSIONS):
        return False
    return True


def score_candidate(url: str, anchor_text: str, priority_keywords: list[str]) -> int:
    normalized_url = url.lower()
    normalized_anchor = anchor_text.lower()
    score = 1

    if is_pdf_url(url):
        score += 10

    for keyword in [*GLOBAL_PRIORITY_KEYWORDS, *priority_keywords]:
        if keyword.lower() in normalized_url:
            score += 6
        if keyword.lower() in normalized_anchor:
            score += 4

    if any(marker in normalized_url for marker in ["quy-che", "quy-dinh", "huong-dan", "thong-bao"]):
        score += 5

    if any(marker in normalized_anchor for marker in ["quy chế", "quy định", "hướng dẫn", "thông báo", "học phí", "học bổng"]):
        score += 5

    path_depth = len([segment for segment in urlparse(url).path.split("/") if segment])
    if 2 <= path_depth <= 6:
        score += 2

    return score


def extract_candidate_links(base_url: str, domain: str, html: str, priority_keywords: list[str]) -> list[UrlCandidate]:
    soup = BeautifulSoup(html, "lxml")
    discovered: dict[str, UrlCandidate] = {}
    for anchor in soup.select("a[href]"):
        href = (anchor.get("href") or "").strip()
        if not href:
            continue

        absolute_url = normalize_url(urljoin(base_url, href))
        if not is_allowed_url(absolute_url, domain):
            continue

        anchor_text = clean_text(anchor.get_text(" ", strip=True))
        score = score_candidate(absolute_url, anchor_text, priority_keywords)
        kind = "pdf" if is_pdf_url(absolute_url) else "html"
        existing = discovered.get(absolute_url)
        if existing is None or score > existing.score:
            discovered[absolute_url] = UrlCandidate(url=absolute_url, score=score, kind=kind)

    return sorted(discovered.values(), key=lambda item: (-item.score, item.url))


def extract_title_and_text(url: str, html: str) -> tuple[str, str]:
    soup = BeautifulSoup(html, "lxml")
    for tag in soup(["script", "style", "noscript", "nav", "header", "footer", "aside", "form"]):
        tag.decompose()

    title = (
        (soup.find("meta", attrs={"property": "og:title"}) or {}).get("content")
        or (soup.find("meta", attrs={"name": "title"}) or {}).get("content")
        or (soup.title.string.strip() if soup.title and soup.title.string else "")
        or (soup.find("h1").get_text(" ", strip=True) if soup.find("h1") else "")
        or urlparse(url).path.strip("/").split("/")[-1]
        or "UIT document"
    )

    main_selectors = [
        "article",
        "main",
        ".node",
        ".content",
        ".entry-content",
        ".post-content",
        "#content",
    ]

    text_candidates: list[str] = []
    for selector in main_selectors:
        for node in soup.select(selector):
            text = clean_text(node.get_text(" ", strip=True))
            if len(text) > 240:
                text_candidates.append(text)

    body_text = clean_text(soup.get_text(" ", strip=True))
    text = max(text_candidates or [body_text], key=len)
    return title.strip(), refine_extracted_html_text(title.strip(), text)


def refine_extracted_html_text(title: str, text: str) -> str:
    cleaned = clean_text(text)
    normalized_title = clean_text(title)

    if normalized_title:
        first_index = cleaned.find(normalized_title)
        if first_index >= 0:
            second_index = cleaned.find(normalized_title, first_index + len(normalized_title))
            if second_index >= 0 and second_index - first_index < 2400:
                cleaned = cleaned[second_index:]
            elif first_index > 500:
                cleaned = cleaned[first_index:]

    for marker in HTML_TAIL_MARKERS:
        marker_index = cleaned.find(marker)
        if marker_index > 180:
            cleaned = cleaned[:marker_index]

    return clean_text(cleaned)


def extract_pdf_text_with_pdftotext(pdf_path: Path) -> str:
    if shutil.which("pdftotext") is None:
        return ""

    command = ["pdftotext", "-layout", str(pdf_path), "-"]
    result = subprocess.run(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        return ""
    return clean_text(result.stdout)


def extract_pdf_text_with_pypdf(pdf_path: Path) -> str:
    try:
        from pypdf import PdfReader
    except Exception:
        return ""

    try:
        reader = PdfReader(str(pdf_path))
    except Exception:
        return ""

    pages: list[str] = []
    for page in reader.pages[:40]:
        try:
            pages.append(page.extract_text() or "")
        except Exception:
            continue
    return clean_text(" ".join(pages))


def extract_pdf_text_with_ocr(pdf_path: Path, work_dir: Path, max_pages: int = 8) -> tuple[str, bool]:
    try:
        from rapidocr_onnxruntime import RapidOCR
    except Exception:
        return "", False

    if shutil.which("pdftoppm") is None:
        return "", False

    image_dir = work_dir / f"{pdf_path.stem}-ocr"
    image_dir.mkdir(parents=True, exist_ok=True)
    image_prefix = image_dir / "page"
    command = [
        "pdftoppm",
        "-png",
        "-f",
        "1",
        "-l",
        str(max_pages),
        str(pdf_path),
        str(image_prefix),
    ]
    result = subprocess.run(command, capture_output=True, text=True, check=False)
    if result.returncode != 0:
        return "", False

    engine = RapidOCR()
    extracted_pages: list[str] = []
    for image_path in sorted(image_dir.glob("*.png")):
        try:
            output, _ = engine(str(image_path))
        except Exception:
            continue
        if not output:
            continue
        page_text = " ".join(item[1] for item in output if len(item) > 1 and item[1])
        if page_text:
            extracted_pages.append(page_text)

    return clean_text(" ".join(extracted_pages)), bool(extracted_pages)


def extract_pdf_text(pdf_path: Path, work_dir: Path) -> tuple[str, bool]:
    text = extract_pdf_text_with_pdftotext(pdf_path)
    if len(text) >= 240:
        return text, False

    fallback_text = extract_pdf_text_with_pypdf(pdf_path)
    if len(fallback_text) > len(text):
        text = fallback_text
    if len(text) >= 240:
        return text, False

    ocr_text, used_ocr = extract_pdf_text_with_ocr(pdf_path, work_dir)
    if len(ocr_text) > len(text):
        text = ocr_text
    return text, used_ocr


def extract_csv_text(file_path: Path) -> str:
    raw_bytes = file_path.read_bytes()
    decoded = ""
    for encoding in ["utf-8-sig", "utf-8", "latin-1"]:
        try:
            decoded = raw_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if not decoded:
        return ""

    reader = csv.reader(StringIO(decoded))
    rows: list[str] = []
    for index, row in enumerate(reader):
        values = [str(value).strip() for value in row if str(value).strip()]
        if not values:
            continue
        rows.append(" | ".join(values))
        if index >= 120:
            break
    return clean_text(" ".join(rows))


def extract_xlsx_text(file_path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return ""

    chunks: list[str] = []
    for sheet in workbook.worksheets[:6]:
        chunks.append(f"Sheet {sheet.title}")
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = [str(value).strip() for value in row if value not in {None, ""}]
            if values:
                chunks.append(" | ".join(values))
            if row_index >= 120:
                break
    workbook.close()
    return clean_text(" ".join(chunks))


def extract_spreadsheet_text(file_path: Path) -> str:
    suffix = file_path.suffix.lower()
    if suffix == ".csv":
        return extract_csv_text(file_path)
    if suffix == ".xlsx":
        return extract_xlsx_text(file_path)
    return ""


def extract_image_text(image_path: Path) -> tuple[str, bool]:
    try:
        from rapidocr_onnxruntime import RapidOCR
    except Exception:
        return "", False

    engine = RapidOCR()
    try:
        output, _ = engine(str(image_path))
    except Exception:
        return "", False
    if not output:
        return "", False
    text = clean_text(" ".join(item[1] for item in output if len(item) > 1 and item[1]))
    return text, bool(text)


def detect_file_type(url: str, content_type: str) -> str:
    normalized_content_type = content_type.lower()
    if is_pdf_url(url) or "application/pdf" in normalized_content_type:
        return "pdf"
    if is_spreadsheet_url(url) or any(
        marker in normalized_content_type
        for marker in ["text/csv", "spreadsheetml", "ms-excel", "application/vnd.ms-excel"]
    ):
        return "spreadsheet"
    if is_image_url(url) or normalized_content_type.startswith("image/"):
        return "image"
    return "html"


async def fetch_sitemap_urls(client: httpx.AsyncClient, source: dict) -> list[UrlCandidate]:
    sitemap_urls: list[str] = []
    robots_url = urljoin(source["base_url"], "/robots.txt")
    try:
        response = await client.get(robots_url)
        if response.status_code == 200:
            for line in response.text.splitlines():
                if line.lower().startswith("sitemap:"):
                    sitemap_urls.append(line.split(":", 1)[1].strip())
    except Exception:
        pass

    for fallback in ["/sitemap.xml", "/sitemap_index.xml"]:
        sitemap_urls.append(urljoin(source["base_url"], fallback))

    unique_sitemaps = list(dict.fromkeys(sitemap_urls))
    discovered: dict[str, UrlCandidate] = {}

    for sitemap_url in unique_sitemaps[:4]:
        try:
            response = await client.get(sitemap_url)
            if response.status_code != 200:
                continue
        except Exception:
            continue

        locs = re.findall(r"<loc>(.*?)</loc>", response.text, flags=re.IGNORECASE)
        for loc in locs[:400]:
            normalized_url = normalize_url(loc)
            if not is_allowed_url(normalized_url, source["domain"]):
                continue
            score = score_candidate(normalized_url, "", source.get("priority_keywords", []))
            kind = "pdf" if is_pdf_url(normalized_url) else "html"
            existing = discovered.get(normalized_url)
            if existing is None or score > existing.score:
                discovered[normalized_url] = UrlCandidate(url=normalized_url, score=score + 3, kind=kind)

    return sorted(discovered.values(), key=lambda item: (-item.score, item.url))


def build_seed_candidates(source: dict) -> list[UrlCandidate]:
    candidates: list[UrlCandidate] = []
    for path in SOURCE_SEED_PATHS.get(source["id"], []):
        absolute_url = normalize_url(urljoin(source["base_url"], path))
        if not is_allowed_url(absolute_url, source["domain"]):
            continue
        candidates.append(
            UrlCandidate(
                url=absolute_url,
                score=score_candidate(absolute_url, "", source.get("priority_keywords", [])) + 20,
                kind="pdf" if is_pdf_url(absolute_url) else "html",
            )
        )
    for absolute_url in SOURCE_DIRECT_URLS.get(source["id"], []):
        normalized_url = normalize_url(absolute_url)
        if not is_allowed_url(normalized_url, source["domain"]):
            continue
        candidates.append(
            UrlCandidate(
                url=normalized_url,
                score=score_candidate(normalized_url, "", source.get("priority_keywords", [])) + 40,
                kind="pdf" if is_pdf_url(normalized_url) else "html",
            )
        )
    if not candidates:
        candidates.append(UrlCandidate(url=normalize_url(source["base_url"]), score=20, kind="html"))
    return candidates


def should_store_document(url: str, title: str, file_type: str) -> bool:
    parsed = urlparse(url)
    normalized_path = parsed.path.rstrip("/")
    normalized_url = url.lower()
    normalized_title = title.lower().strip()
    match_text = re.sub(r"[^0-9a-zà-ỹ]+", " ", f"{normalized_url} {normalized_title}".lower())

    if normalized_path in SKIP_STORE_PATHS:
        return False
    if "/category/" in normalized_path:
        return False
    if re.search(r"/page/\d+$", normalized_path):
        return False
    if normalized_title in SKIP_STORE_TITLES:
        return False
    if any(marker in normalized_title for marker in LOW_VALUE_TITLE_MARKERS):
        return False
    if "2-uncategorised" in normalized_url:
        uncategorized_tail = normalized_url.split("2-uncategorised", maxsplit=1)[-1]
        uncategorized_match_text = re.sub(r"[^0-9a-zà-ỹ]+", " ", f"{uncategorized_tail} {normalized_title}".lower())
        if any(marker in uncategorized_match_text for marker in {"module variations", "icons", "logo", "gioi thieu", "giới thiệu", "tuyen sinh", "tuyển sinh", "lien thong", "liên thông", "cao hoc", "cao học", "tuyen dung", "tuyển dụng"}):
            return False
        if any(marker in uncategorized_tail or marker in normalized_title for marker in UNCATEGORIZED_BLOCK_MARKERS):
            return False
        if not any(
            marker in uncategorized_match_text
            for marker in {
                "chuan dau ra",
                "chuẩn đầu ra",
                "mon hoc",
                "môn học",
                "hoc phan",
                "học phần",
                "de cuong",
                "đề cương",
                "khoa luan",
                "khóa luận",
                "kltn",
                "tot nghiep",
                "tốt nghiệp",
                "thuc tap",
                "thực tập",
                "internship",
                "hoc vu",
                "học vụ",
                "hoc bong",
                "học bổng",
                "doan hoi",
                "đoàn hội",
                "cau lac bo",
                "câu lạc bộ",
                "clb",
                "nghien cuu",
                "nghiên cứu",
            }
        ):
            return False

    if "tuyen-dung" in normalized_url or "tuyen dung" in match_text or "tuyển dụng" in match_text:
        return False
    if "tuyen-sinh" in normalized_url or "tuyen sinh" in match_text or "tuyển sinh" in match_text:
        return False
    if "lienthong" in normalized_url or "lien thong" in match_text or "liên thông" in match_text:
        return False
    if "cao-hoc" in normalized_url or "cao hoc" in match_text or "cao học" in match_text:
        return False
    if normalized_title in {
        "trường đại học công nghệ thông tin - đại học quốc gia thành phố hồ chí minh",
        "trường đại học công nghệ thông tin - đại học quốc gia thành phố hồ chí minh | cổng thông tin đào tạo",
    }:
        return False

    if file_type == "pdf":
        return True
    return True


def load_registry(allowed_domains: set[str] | None = None) -> list[dict]:
    data_dir = resolve_data_dir()
    registry_path = data_dir / "source_registry.json"
    registry = json.loads(registry_path.read_text(encoding="utf-8"))
    filtered_registry = [source for source in registry if source.get("id") not in TRAINING_DISABLED_SOURCE_IDS]
    if allowed_domains is None:
        return filtered_registry
    return [source for source in filtered_registry if source.get("domain") in allowed_domains]


def load_curated_overrides(data_dir: Path) -> dict[str, dict]:
    curated_dir = data_dir / "curated"
    manifest_path = curated_dir / "manifest.json"
    if not manifest_path.exists():
        return {}

    overrides: dict[str, dict] = {}
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    for item in manifest:
        relative_file = item.get("file")
        if not relative_file:
            continue
        content_path = curated_dir / relative_file
        if not content_path.exists():
            continue
        text = clean_text(content_path.read_text(encoding="utf-8"))
        if not text:
            continue
        overrides[normalize_url(str(item["url"]))] = {
            "title": item.get("title"),
            "summary": item.get("summary"),
            "tags": item.get("tags", []),
            "text": text,
        }
    return overrides


def source_target_limit(source: dict, per_source_target: int) -> int:
    max_pages = int(source.get("max_pages", 4))
    source_id = str(source.get("id", ""))
    if source_id in STUDENT_CRITICAL_SOURCE_IDS:
        return min(max(12, max_pages * 6), per_source_target + 12)
    if source_id in FACULTY_SOURCE_IDS:
        return max(4, min(max_pages * 3, max(6, per_source_target // 2)))
    return min(max(10, max_pages * 5), per_source_target)


def make_snapshot_markdown(title: str, source_name: str, url: str, crawled_at: str, text: str, file_type: str) -> str:
    return "\n".join(
        [
            f"# {title}",
            "",
            f"- URL: {url}",
            f"- Source: {source_name}",
            f"- Crawled at: {crawled_at}",
            f"- File type: {file_type}",
            "",
            "## Extracted content",
            "",
            text,
            "",
        ]
    )


def write_jsonl(path: Path, records: list[dict]) -> None:
    with path.open("w", encoding="utf-8") as file_pointer:
        for record in records:
            file_pointer.write(json.dumps(record, ensure_ascii=False) + "\n")


async def collect_training_corpus(
    target_documents: int = 30,
    min_text_length: int = 240,
    allowed_domains: set[str] | None = None,
) -> dict:
    data_dir = resolve_data_dir()
    snapshots_dir = data_dir / "snapshots"
    processed_dir = data_dir / "processed"
    pdf_dir = data_dir / "pdfs"
    ocr_dir = data_dir / "ocr"
    for directory in [snapshots_dir, processed_dir, pdf_dir, ocr_dir]:
        directory.mkdir(parents=True, exist_ok=True)

    registry = load_registry(allowed_domains=allowed_domains)
    curated_overrides = load_curated_overrides(data_dir)
    records: list[dict] = []
    seen_urls: set[str] = set()
    seen_hashes: set[str] = set()
    total_target = max(target_documents, 1)
    total_sources = max(len(registry), 1)
    # Keep the crawl balanced across UIT domains so the corpus does not get
    # dominated by the first sources in the registry order.
    per_source_target = max(12, ((total_target + total_sources - 1) // total_sources) + 4)

    async with httpx.AsyncClient(
        timeout=httpx.Timeout(25.0, connect=10.0),
        headers={"User-Agent": "StudifyTrainer/1.0 (+https://uit.edu.vn)"},
        follow_redirects=True,
    ) as client:
        for source in registry:
            source_records: list[dict] = []
            source_snapshot_dir = snapshots_dir / source["id"]
            source_pdf_dir = pdf_dir / source["id"]
            source_snapshot_dir.mkdir(parents=True, exist_ok=True)
            source_pdf_dir.mkdir(parents=True, exist_ok=True)
            source_target = source_target_limit(source, per_source_target)
            discovery_limit = max(40, int(source.get("max_pages", 4)) * 14)

            queue: dict[str, UrlCandidate] = {}
            visited_urls: set[str] = set()
            for candidate in build_seed_candidates(source):
                queue[candidate.url] = candidate

            for candidate in (await fetch_sitemap_urls(client, source))[:40]:
                existing = queue.get(candidate.url)
                if existing is None or candidate.score > existing.score:
                    queue[candidate.url] = candidate

            print(f"[train] source={source['id']} queued={len(queue)}")

            while queue and len(source_records) < source_target:
                candidate = sorted(queue.values(), key=lambda item: (-item.score, item.url))[0]
                queue.pop(candidate.url, None)

                if candidate.url in visited_urls or candidate.url in seen_urls:
                    continue
                visited_urls.add(candidate.url)

                try:
                    response = await client.get(candidate.url)
                    response.raise_for_status()
                except Exception as exc:
                    print(f"[skip] {candidate.url} -> {exc}")
                    continue

                content_type = response.headers.get("content-type", "")
                file_type = detect_file_type(candidate.url, content_type)
                override = curated_overrides.get(candidate.url)

                if file_type in {"pdf", "spreadsheet", "image"}:
                    title = Path(urlparse(candidate.url).path).name or "uit-pdf"
                    file_extension = Path(urlparse(candidate.url).path).suffix.lower()
                    if not file_extension:
                        file_extension = ".pdf" if file_type == "pdf" else ".bin"
                    file_name = f"{len(source_records) + 1:02d}-{slugify(title)}{file_extension}"
                    download_path = source_pdf_dir / file_name
                    download_path.write_bytes(response.content)

                    used_ocr = False
                    if file_type == "pdf":
                        extracted_text, used_ocr = extract_pdf_text(download_path, ocr_dir)
                    elif file_type == "spreadsheet":
                        extracted_text = extract_spreadsheet_text(download_path)
                    else:
                        extracted_text, used_ocr = extract_image_text(download_path)

                    if override:
                        title = str(override.get("title") or title)
                        extracted_text = str(override.get("text") or extracted_text)

                    if len(extracted_text) < min_text_length:
                        print(f"[skip-{file_type}] short text {candidate.url}")
                        continue
                    if not should_store_document(candidate.url, title, file_type):
                        continue

                    hash_value = content_hash(extracted_text)
                    if hash_value in seen_hashes:
                        continue

                    seen_hashes.add(hash_value)
                    seen_urls.add(candidate.url)
                    crawled_at = now_iso()
                    snapshot_name = f"{len(source_records) + 1:02d}-{slugify(title)}.md"
                    snapshot_path = source_snapshot_dir / snapshot_name
                    snapshot_path.write_text(
                        make_snapshot_markdown(title, source["name"], candidate.url, crawled_at, extracted_text, file_type),
                        encoding="utf-8",
                    )

                    record = {
                        "source_id": source["id"],
                        "source_name": source["name"],
                        "base_url": source["base_url"],
                        "domain": source["domain"],
                        "is_official_uit": bool(source["official"]),
                        "title": title,
                        "url": candidate.url,
                        "snapshot_file": str(snapshot_path.relative_to(data_dir)),
                        "download_file": str(download_path.relative_to(data_dir)),
                        "text": extracted_text,
                        "summary": str(override.get("summary") or extracted_text[:500]) if override else extracted_text[:500],
                        "crawled_at": crawled_at,
                        "tags": list(dict.fromkeys([*source.get("priority_keywords", []), *(override.get("tags", []) if override else [])])),
                        "file_type": file_type,
                        "ocr_used": used_ocr,
                    }
                    source_records.append(record)
                    records.append(record)
                    print(f"[{file_type}] {candidate.url}")
                    continue

                html = response.text
                page_title, page_text = extract_title_and_text(candidate.url, html)
                if override:
                    page_title = str(override.get("title") or page_title)
                    page_text = str(override.get("text") or page_text)
                if len(page_text) >= min_text_length and should_store_document(candidate.url, page_title, "html"):
                    hash_value = content_hash(page_text)
                    if hash_value not in seen_hashes:
                        seen_hashes.add(hash_value)
                        seen_urls.add(candidate.url)
                        crawled_at = now_iso()
                        snapshot_name = f"{len(source_records) + 1:02d}-{slugify(page_title)}.md"
                        snapshot_path = source_snapshot_dir / snapshot_name
                        snapshot_path.write_text(
                            make_snapshot_markdown(page_title, source["name"], candidate.url, crawled_at, page_text, "html"),
                            encoding="utf-8",
                        )
                        record = {
                            "source_id": source["id"],
                            "source_name": source["name"],
                            "base_url": source["base_url"],
                            "domain": source["domain"],
                            "is_official_uit": bool(source["official"]),
                            "title": page_title,
                            "url": candidate.url,
                            "snapshot_file": str(snapshot_path.relative_to(data_dir)),
                            "text": page_text,
                            "summary": str(override.get("summary") or page_text[:500]) if override else page_text[:500],
                            "crawled_at": crawled_at,
                            "tags": list(dict.fromkeys([*source.get("priority_keywords", []), *(override.get("tags", []) if override else [])])),
                            "file_type": "html",
                            "ocr_used": False,
                        }
                        source_records.append(record)
                        records.append(record)
                        print(f"[html] {candidate.url}")

                for discovered in extract_candidate_links(candidate.url, source["domain"], html, source.get("priority_keywords", []))[:discovery_limit]:
                    if discovered.url in visited_urls or discovered.url in seen_urls:
                        continue
                    existing = queue.get(discovered.url)
                    if existing is None or discovered.score > existing.score:
                        queue[discovered.url] = discovered

            manifest_path = source_snapshot_dir / "manifest.json"
            manifest_path.write_text(json.dumps(source_records, ensure_ascii=False, indent=2), encoding="utf-8")

    manifest_path = processed_dir / "manifest.json"
    jsonl_path = processed_dir / "uit_documents.jsonl"
    manifest_path.write_text(json.dumps(records, ensure_ascii=False, indent=2), encoding="utf-8")
    write_jsonl(jsonl_path, records)

    summary = {
        "target_documents": total_target,
        "collected_documents": len(records),
        "processed_jsonl": str(jsonl_path),
        "manifest": str(manifest_path),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Prepare the Studify RAG corpus before Docker startup.")
    parser.add_argument("--target-documents", type=int, default=280, help="Target number of documents to collect.")
    parser.add_argument("--min-text-length", type=int, default=240, help="Minimum cleaned text length for a kept document.")
    parser.add_argument(
        "--domains",
        type=str,
        default="",
        help="Comma-separated list of domains to crawl, for example daa.uit.edu.vn,student.uit.edu.vn,khtc.uit.edu.vn",
    )
    parser.add_argument(
        "--student-focus",
        action="store_true",
        help="Only crawl the student-critical official sources: UIT, DAA, Student, CTSV, KHTC và OEP.",
    )
    return parser.parse_args()


async def main() -> None:
    args = parse_args()
    allowed_domains: set[str] | None = None
    if args.student_focus:
        allowed_domains = {
            "uit.edu.vn",
            "daa.uit.edu.vn",
            "student.uit.edu.vn",
            "ctsv.uit.edu.vn",
            "khtc.uit.edu.vn",
            "oep.uit.edu.vn",
        }
    if args.domains.strip():
        allowed_domains = {item.strip() for item in args.domains.split(",") if item.strip()}
    await collect_training_corpus(
        target_documents=args.target_documents,
        min_text_length=args.min_text_length,
        allowed_domains=allowed_domains,
    )


if __name__ == "__main__":
    asyncio.run(main())
