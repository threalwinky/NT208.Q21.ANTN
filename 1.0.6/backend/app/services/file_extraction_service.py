from __future__ import annotations

import csv
import shutil
import subprocess
from io import StringIO
from pathlib import Path

from app.services.text_utils import clean_text


class FileExtractionService:
    def extract_text_from_path(self, file_path: Path) -> tuple[str, bool, str]:
        suffix = file_path.suffix.lower()
        if suffix == ".pdf":
            text, used_ocr = self.extract_pdf_text(file_path, file_path.parent / "ocr")
            return text, used_ocr, "pdf"
        if suffix in {".png", ".jpg", ".jpeg"}:
            text, used_ocr = self.extract_image_text(file_path)
            return text, used_ocr, "image"
        if suffix in {".csv", ".xlsx"}:
            return self.extract_spreadsheet_text(file_path), False, "spreadsheet"
        if suffix in {".txt", ".md"}:
            return clean_text(file_path.read_text(encoding="utf-8", errors="ignore")), False, "text"
        return clean_text(file_path.read_text(encoding="utf-8", errors="ignore")), False, "text"

    def extract_pdf_text(self, pdf_path: Path, work_dir: Path) -> tuple[str, bool]:
        text = self._extract_pdf_text_with_pdftotext(pdf_path)
        if len(text) >= 240:
            return text, False

        fallback_text = self._extract_pdf_text_with_pypdf(pdf_path)
        if len(fallback_text) > len(text):
            text = fallback_text
        if len(text) >= 240:
            return text, False

        ocr_text, used_ocr = self._extract_pdf_text_with_ocr(pdf_path, work_dir)
        if len(ocr_text) > len(text):
            text = ocr_text
        return text, used_ocr

    def _extract_pdf_text_with_pdftotext(self, pdf_path: Path) -> str:
        if shutil.which("pdftotext") is None:
            return ""
        command = ["pdftotext", "-layout", str(pdf_path), "-"]
        result = subprocess.run(command, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            return ""
        return clean_text(result.stdout)

    def _extract_pdf_text_with_pypdf(self, pdf_path: Path) -> str:
        try:
            from pypdf import PdfReader
        except Exception:
            return ""

        try:
            reader = PdfReader(str(pdf_path))
        except Exception:
            return ""

        pages: list[str] = []
        for page in reader.pages[:40]:
            try:
                pages.append(page.extract_text() or "")
            except Exception:
                continue
        return clean_text(" ".join(pages))

    def _extract_pdf_text_with_ocr(self, pdf_path: Path, work_dir: Path, max_pages: int = 8) -> tuple[str, bool]:
        try:
            from rapidocr_onnxruntime import RapidOCR
        except Exception:
            return "", False

        if shutil.which("pdftoppm") is None:
            return "", False

        image_dir = work_dir / f"{pdf_path.stem}-ocr"
        image_dir.mkdir(parents=True, exist_ok=True)
        image_prefix = image_dir / "page"
        command = [
            "pdftoppm",
            "-png",
            "-f",
            "1",
            "-l",
            str(max_pages),
            str(pdf_path),
            str(image_prefix),
        ]
        result = subprocess.run(command, capture_output=True, text=True, check=False)
        if result.returncode != 0:
            return "", False

        engine = RapidOCR()
        extracted_pages: list[str] = []
        for image_path in sorted(image_dir.glob("*.png")):
            try:
                output, _ = engine(str(image_path))
            except Exception:
                continue
            if not output:
                continue
            page_text = " ".join(item[1] for item in output if len(item) > 1 and item[1])
            if page_text:
                extracted_pages.append(page_text)
        return clean_text(" ".join(extracted_pages)), bool(extracted_pages)

    def extract_spreadsheet_text(self, file_path: Path) -> str:
        suffix = file_path.suffix.lower()
        if suffix == ".csv":
            return self._extract_csv_text(file_path)
        if suffix == ".xlsx":
            return self._extract_xlsx_text(file_path)
        return ""

    def _extract_csv_text(self, file_path: Path) -> str:
        raw_bytes = file_path.read_bytes()
        decoded = ""
        for encoding in ["utf-8-sig", "utf-8", "latin-1"]:
            try:
                decoded = raw_bytes.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        if not decoded:
            return ""

        reader = csv.reader(StringIO(decoded))
        rows: list[str] = []
        for index, row in enumerate(reader):
            values = [str(value).strip() for value in row if str(value).strip()]
            if not values:
                continue
            rows.append(" | ".join(values))
            if index >= 120:
                break
        return clean_text(" ".join(rows))

    def _extract_xlsx_text(self, file_path: Path) -> str:
        try:
            from openpyxl import load_workbook
        except Exception:
            return ""

        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
        except Exception:
            return ""

        chunks: list[str] = []
        for sheet in workbook.worksheets[:6]:
            chunks.append(f"Sheet {sheet.title}")
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(value).strip() for value in row if value not in {None, ""}]
                if values:
                    chunks.append(" | ".join(values))
                if row_index >= 120:
                    break
        workbook.close()
        return clean_text(" ".join(chunks))

    def extract_image_text(self, image_path: Path) -> tuple[str, bool]:
        try:
            from rapidocr_onnxruntime import RapidOCR
        except Exception:
            return "", False

        engine = RapidOCR()
        try:
            output, _ = engine(str(image_path))
        except Exception:
            return "", False
        if not output:
            return "", False
        text = clean_text(" ".join(item[1] for item in output if len(item) > 1 and item[1]))
        return text, bool(text)
